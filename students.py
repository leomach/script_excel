import os
import json
import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from cities import remove_duplicate_keys, cities
from exit_process.core import (
    DATAFRAMES_STUDENTS,
    DATAFRAMES_TEAMS,
)


# Remove chaves duplicadas
clean_cities = remove_duplicate_keys(cities)

# Salva no arquivo JSON
with open("clean_cities.json", "w") as file:
    json.dump(clean_cities, file, indent=4, ensure_ascii=False)

print("Arquivo JSON salvo com sucesso!")

# Definir o caminho onde os arquivos serão salvos
global caminho_saida
caminho_saida = "./saida/"
# Garantir que o diretório existe
os.makedirs(caminho_saida, exist_ok=True)



def professions(df_pais, escola):
    global caminho_saida
    # Extrair as colunas das profissões
    df_profession = df_pais[['_profession', 'id']].copy()

    # Renomear as colunas para facilitar o processamento
    df_profession.rename(columns={'id': 'id_pais', '_profession': 'description'}, inplace=True)

    # Remover espaços no início e no final das strings, e eliminar espaços duplicados nas colunas relevantes
    df_profession['description'] = df_profession['description'].str.strip().str.replace(r'\s+', ' ', regex=True)

    # Substituir NaN por uma string vazia (caso queira ignorar as células NaN)
    df_profession.replace({'': pd.NA, 'Não Informado': pd.NA}, inplace=True)

    df_profession.dropna(subset=['description'], inplace=True)

    # Agrupar os ids de responsaveis de profissões duplicados separando-os por vírgula
    df_profession = df_profession.groupby(['description'], as_index=False).agg({
        'id_pais': lambda x: ', '.join(x.astype(str).dropna())  # Concatena as IDs de responsaveis
    })

    df_profession['profession'] = range(1, len(df_profession) + 1)

    # Explodir IDs corretamente
    df_profession_exploded = df_profession.copy().assign(
        id_pais=df_profession['id_pais'].astype(str).str.split(', ')
    ).explode('id_pais')

    # Converter IDs para int, substituindo NaN por 0
    df_profession_exploded['id_pais'] = df_profession_exploded['id_pais'].fillna(0).astype(int)
    df_profession_exploded['profession'] = df_profession_exploded['profession'].fillna(0).astype(int)


    # Fazer merge novamente garantindo que todos os estudantes tenham mãe/pai
    df_responsaveis_com_profissoes = pd.merge(df_pais, df_profession_exploded[['id_pais', 'profession']], left_on='id', right_on='id_pais', how='left')
    df_responsaveis_com_prof = df_responsaveis_com_profissoes[['name', 'sex', 'profession', 'id']].copy()

    # Preparar para impressão
    df_profession_print = df_profession[['profession', 'description']].copy()
    df_profession_print.rename(columns={'profession': 'id',}, inplace=True)

    # Salvar o DataFrame como um arquivo CSV
    df_profession_print.to_csv(f"{caminho_saida}profissoes-{escola}.csv", index=False, encoding="utf-8-sig", sep=";")

    # Exibir as primeiras linhas para verificar
    # print("###################### df_profession #####################")
    # print(df_profession.head())

    return df_responsaveis_com_prof


def responsibles(df_estudantes, escola):
    global caminho_saida
    # Extrair as colunas dos pais
    df_pais = df_estudantes[['*Filiação1', 'ProfFiliação1', 'Filiação2', 'ProfFiliação2', 'id']].copy()

    # Renomear as colunas para facilitar o processamento
    df_pais.rename(columns={'*Filiação1': '1', 'ProfFiliação1': 'Profissao_Mae', 'Filiação2': '2', 'ProfFiliação2': 'Profissao_Pai', 'id': 'id_student'}, inplace=True)

    # Remover espaços no início e no final das strings, e eliminar espaços duplicados nas colunas relevantes
    df_pais['1'] = df_pais['1'].str.strip().str.replace(r'\s+', ' ', regex=True)
    df_pais['2'] = df_pais['2'].str.strip().str.replace(r'\s+', ' ', regex=True)
    df_pais['Profissao_Mae'] = df_pais['Profissao_Mae'].str.strip().str.replace(r'\s+', ' ', regex=True)
    df_pais['Profissao_Pai'] = df_pais['Profissao_Pai'].str.strip().str.replace(r'\s+', ' ', regex=True)

    # Substituir NaN por uma string vazia (caso queira ignorar as células NaN)
    df_pais['id_student'] = df_pais['id_student'].fillna('')
    df_pais['1'] = df_pais['1'].fillna('')
    df_pais['2'] = df_pais['2'].fillna('')
    df_pais['Profissao_Mae'] = df_pais['Profissao_Mae'].fillna('Não Informado')
    df_pais['Profissao_Pai'] = df_pais['Profissao_Pai'].fillna('Não Informado')

    # Agrupar os ids de estudantes de responsáveis duplicados separando-os por vírgula
    df_pais = df_pais.groupby(['1', '2', 'Profissao_Mae', 'Profissao_Pai'], as_index=False).agg({
        'id_student': lambda x: ', '.join(x.astype(str).dropna())  # Concatena as IDs de estudantes
    })

    # Separar pais e mães
    df_pai = df_pais[["2", "Profissao_Pai", "id_student"]].copy()
    df_mae = df_pais[["1", "Profissao_Mae", "id_student"]].copy()

    # Ajustar o DataFrame da mãe
    df_mae["name"] = df_mae["1"]
    df_mae["sex"] = "1"  # Mãe será associada ao '1'
    df_mae["_profession"] = df_mae["Profissao_Mae"]
    df_mae = df_mae[["name", "sex", "_profession", "id_student"]]
    df_mae['mother'] = range(1, len(df_mae) + 1)

    # Ajustar o DataFrame do pai
    df_pai["name"] = df_pai["2"]
    df_pai["sex"] = "2"  # Pai será associado ao '2'
    df_pai["_profession"] = df_pai["Profissao_Pai"]
    df_pai = df_pai[["name", "sex", "_profession", "id_student"]]
    # Criar os ids para o pai, continuando a partir do último id da mãe
    df_pai['father'] = range(len(df_mae) + 1, len(df_mae) + 1 + len(df_pai))
    
    # Remover pais/mães sem nome
    df_pai.dropna(subset=['name'], inplace=True)
    df_mae.dropna(subset=['name'], inplace=True)

    # Criar os IDs dos pais e mães
    df_mae['mother'] = range(1, len(df_mae) + 1)
    df_pai['father'] = range(len(df_mae) + 1, len(df_mae) + 1 + len(df_pai))

    # Explodir IDs corretamente
    df_pais_exploded = df_pai.copy().assign(
        id_student=df_pai['id_student'].astype(str).str.split(', ')
    ).explode('id_student')
    
    df_maes_exploded = df_mae.copy().assign(
        id_student=df_mae['id_student'].astype(str).str.split(', ')
    ).explode('id_student')
    

    # Converter IDs para int, substituindo NaN por 0
    df_pais_exploded['id_student'] = df_pais_exploded['id_student'].fillna(0).astype(int)
    df_maes_exploded['id_student'] = df_maes_exploded['id_student'].fillna(0).astype(int)

    df_pais_exploded['father'] = df_pais_exploded['father'].fillna(0).astype(int)
    df_maes_exploded['mother'] = df_maes_exploded['mother'].fillna(0).astype(int)


    # Fazer merge novamente garantindo que todos os estudantes tenham mãe/pai
    df_estudantes_com_pais = pd.merge(df_estudantes, df_pais_exploded[['father', 'id_student']], left_on='id', right_on='id_student', how='left')
    df_estudantes_com_responsaveis = pd.merge(df_estudantes_com_pais, df_maes_exploded[['mother', 'id_student']], left_on='id', right_on='id_student', how='left')

    # Concatenar os dois DataFrames de pai e mãe
    df_pai_rename = df_pai.copy().rename(columns={'father': 'id'})
    df_mae_rename = df_mae.copy().rename(columns={'mother': 'id'})
    df_pais_combined = pd.concat([df_mae_rename, df_pai_rename], ignore_index=True)

    # Remover linhas onde o nome do parente é NaN
    df_pais_combined.dropna(subset=['name'], inplace=True)

    df_responsaveis_com_prof = professions(df_pais=df_pais_combined, escola=escola)
    df_responsaveis_com_prof['profession'] = df_responsaveis_com_prof['profession'].astype('Int64')
    df_responsaveis_com_prof.dropna(subset=['name'], inplace=True)

    # Salvar o DataFrame como um arquivo CSV
    df_responsaveis_com_prof.to_csv(f"{caminho_saida}responsaveis-{escola}.csv", index=False, encoding="utf-8-sig", sep=";")

    # Exibir as primeiras linhas para verificar
    # print("###################### pais #####################")
    # print(df_responsaveis_com_prof.head())

    return df_estudantes_com_responsaveis



def students_archives(escola, arquivo, base_dir):
    global caminho_saida

    # Caminho para um dos arquivos de estudantes (ajuste conforme necessário)
    caminho_arquivo = f"{base_dir}/{escola}/Estudantes/{arquivo}"

    # Ler o arquivo Excel, pulando as primeiras 6 linhas
    df_student = pd.read_excel(caminho_arquivo, engine="openpyxl", skiprows=6)
    df_estudantes = df_student[['*NomeEstudante', '*Dt Nasc.', '*Sexo', "*Filiação1", "ProfFiliação1", "Filiação2",	"ProfFiliação2", "Email", "*Raça/Cor", "*Nacionalidade", "*Município Nasc.", "*PCD", "*CPF", "RG","Dígito RG","Orgão Exp.","UF", "*Endereço","*Número", "*Bairro",	"*CEP",	"*Município",]].copy()

    # Criar uma coluna de ID único para cada estudante
    df_estudantes['id'] = range(1, len(df_estudantes) + 1)

    df_estudantes_com_responsaveis = responsibles(df_estudantes, escola)

    df_estudantes_final = df_estudantes_com_responsaveis[['id','*NomeEstudante', '*Dt Nasc.', '*Sexo', "Email", "*Município Nasc.", "*CPF", "RG","Dígito RG","Orgão Exp.","UF", "*Endereço","*Número", "*Bairro", "*CEP", "*Município", "mother", "father"]].copy()

    # TODO Renomear os nomes das colunas de estudantes
    df_estudantes_com_responsaveis_rename = df_estudantes_final.copy().rename(columns={'*NomeEstudante': 'name', '*Sexo': 'sex', '*Dt Nasc.': 'date_birth', "Email": "email", "*Município Nasc.": "city_birth", '*CPF': 'cpf', "Orgão Exp.": "expedition_organ", "UF": 'state_birth', '*Endereço': 'address', "*Número": "number", "*Bairro": "district", "*CEP": "zip_code", "*Município": "city"})

    df_estudantes_com_responsaveis_rename = df_estudantes_com_responsaveis_rename.fillna("")

    df_estudantes_com_responsaveis_rename['sex'] = df_estudantes_com_responsaveis_rename['sex'].map({'M': 1, 'F': 2}).fillna(df_estudantes_com_responsaveis_rename['sex'])

    df_estudantes_com_responsaveis_rename['rg'] = df_estudantes_com_responsaveis_rename['RG'].astype(str).str.strip() + '-' + df_estudantes_com_responsaveis_rename['Dígito RG'].astype(str).str.strip()

    df_estudantes_com_responsaveis_rename['rg'] = df_estudantes_com_responsaveis_rename[['RG', 'Dígito RG']].astype(str).apply(lambda x: '-'.join(x).strip('-') if x.notna().all() else '', axis=1)

    # Carregar o JSON com os códigos das cidades
    with open("./clean_cities.json", "r", encoding="utf-8") as f:
        city_mapping = json.load(f)[0]  # O JSON parece estar dentro de uma lista

    # Converter as chaves do mapeamento para inteiros
    city_mapping = {int(k): v for k, v in city_mapping.items()}

    # Garantir que os valores da coluna "city" e "city_birth" sejam inteiros antes de mapear
    df_estudantes_com_responsaveis_rename["city"] = (
        pd.to_numeric(df_estudantes_com_responsaveis_rename["city"], errors="coerce")
        .fillna(0)  # Preenche NaN temporariamente para evitar erro na conversão
        .astype(int)
        .map(city_mapping)
    )

    df_estudantes_com_responsaveis_rename["city_birth"] = (
        pd.to_numeric(df_estudantes_com_responsaveis_rename["city_birth"], errors="coerce")
        .fillna(0)  # Preenche NaN temporariamente para evitar erro na conversão
        .astype(int)
        .map(city_mapping)
    )

    df_estudantes_com_responsaveis_rename = df_estudantes_com_responsaveis_rename.drop(columns=["RG","Dígito RG"])


    # Salvar o DataFrame como um arquivo CSV
    df_estudantes_com_responsaveis_rename.to_csv(f"{caminho_saida}alunos-{escola}.csv", index=False, encoding="utf-8-sig", sep=";")

    # Exibir as primeiras linhas para verificar
    # print("###################### Estudantes #####################")
    # print(df_estudantes_com_responsaveis_rename.head(30))

def students_archives_unify(escola, arquivo, base_dir, output_file="./entrada/TODOS OS DADOS/Estudantes/todos_os_alunos.xlsx"):
    global DATAFRAMES_STUDENTS

    # Caminho para um dos arquivos de estudantes (ajuste conforme necessário)
    caminho_arquivo = f"{base_dir}/{escola}/Estudantes/{arquivo}"

    # Ler o arquivo Excel, pulando as primeiras 6 linhas
    df_student = pd.read_excel(caminho_arquivo, engine="openpyxl", skiprows=6)
    df_estudantes = df_student[['*NomeEstudante', '*Dt Nasc.', '*Sexo', "*Filiação1", "ProfFiliação1", "Filiação2",	"ProfFiliação2", "Email", "*Raça/Cor", "*Nacionalidade", "*Município Nasc.", "*PCD", "*CPF", "RG","Dígito RG","Orgão Exp.","UF", "*Endereço","*Número", "*Bairro",	"*CEP",	"*Município",]].copy()

    # Atualiza o DATAFRAMES_STUDENTS para ter todos os estudantes
    if DATAFRAMES_STUDENTS is None or DATAFRAMES_STUDENTS.empty:
        DATAFRAMES_STUDENTS = df_estudantes.copy()

        # Criar o arquivo do zero caso ainda não exista
        df_estudantes.to_excel(output_file, sheet_name="Estudantes", index=False, engine="openpyxl")

        print("Dataframe inicial criado")
    else:
        DATAFRAMES_STUDENTS = pd.concat(
                [DATAFRAMES_STUDENTS, df_estudantes], 
                ignore_index=True, 
                join="outer"  # Inclui todas as colunas existentes
            )
        
        if os.path.exists(output_file):
            # Carregar o arquivo Excel existente
            book = load_workbook(output_file)

            # Verifica se a aba "Estudantes" existe
            if "Estudantes" not in book.sheetnames:
                print("Aba 'Estudantes' não encontrada. Criando nova aba.")
                df_estudantes.to_excel(output_file, sheet_name="Estudantes", index=False, engine="openpyxl")
                return

            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = book["Estudantes"]
                startrow = sheet.max_row if sheet.max_row else 1  # Pega a última linha da aba
                df_estudantes.to_excel(writer, sheet_name="Estudantes", index=False, startrow=startrow, header=False)

    total_linhas = len(DATAFRAMES_STUDENTS)
    print(f"🧑‍🎓 Total de alunos atualizado: {total_linhas}")
    print(f"✅ Dados adicionados a {output_file}")
