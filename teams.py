import os
import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore

from exit_process.core import DATAFRAMES_TEAMS

# Definir o caminho onde os arquivos serão salvos
global caminho_saida
caminho_saida = "./saida/"
# Garantir que o diretório existe
os.makedirs(caminho_saida, exist_ok=True)

def team_archives(escola, arquivo, base_dir):
    global caminho_saida
    # Caminho para um dos arquivos de estudantes (ajuste conforme necessário)
    caminho_arquivo = f"{base_dir}/{escola}/Professores/{arquivo}"

    df_team = pd.read_excel(caminho_arquivo, engine="openpyxl")
    df_teams = df_team[["CPF", "Nome", "Função", "Telefone", "Email", "Observação"]].copy()

    # Dividindo por função
    df_professores = df_teams[df_teams["Função"].isin(["EDUCADOR", "Educador (a)", "Educador(a)", "Mediador", "Mediador(a)", "Mediadora", "Monitora", "PROFESSOR", "Professora"])].copy()
    df_adm = df_teams[df_teams["Função"].isin(["Administrador", "AUX. DE SECRETARIA", "Diretor", "DIRETOR ADJUNTO", "Diretora", "DIRETORA ADJUNTA", "Diretora Adjunta", "GESTOR", "GESTORA", "SECRETARIA", "Secretária", "TEC. SECRETARIA", "Técnica de secretaria", "Técnico Administrativo", "TECNICO DE SECRETARIA", "Técnico de Secretaria", "Técnico(a) de Secretaria"])].copy()
    df_coordenadores = df_teams[df_teams["Função"].isin(["Coodenador(a)", "Coordenador", "Coordenador(a)", "Coordenadora", "Coordenadora Pedagógica",])].copy()

    # Ajustando IDs de função
    df_professores.loc[:, "Função"] = "01"
    df_adm.loc[:, "Função"] = "04"
    df_coordenadores.loc[:, "Função"] = "01"

    df_teams = pd.concat([df_adm, df_coordenadores], ignore_index=True)

    # Ajustando nomes das colunas
    df_teachers = df_professores.copy().rename(columns={"CPF": "cpf", "Nome": "name", "Função": "type", "Telefone": "celphone_1", "Email": "email", "Observação": "note"})
    df_teams = df_teams.rename(columns={"CPF": "cpf", "Nome": "name", "Função": "team_type", "Telefone": "celphone_1", "Email": "email", "Observação": "note"})

    df_teachers['id'] = range(1, len(df_teachers) + 1)
    df_teams['id'] = range(1, len(df_teams) + 1)

    # Adicionando login e password de acordo com o CPF
    df_teachers['login'] = df_teachers['cpf']
    df_teachers['password'] = df_teachers['cpf']
    df_teams['login'] = df_teams['cpf']
    df_teams['password'] = df_teams['cpf']

    # Salvando os DataFrames como arquivos CSV
    df_teachers.to_csv(f"{caminho_saida}professores-{escola}.csv", index=False, encoding="utf-8-sig", sep=";")
    df_teams.to_csv(f"{caminho_saida}equipe-{escola}.csv", index=False, encoding="utf-8-sig", sep=";")

def team_archives_unify(escola, arquivo, base_dir, output_file="./entrada/TODOS OS DADOS/Professores/toda_equipe.xlsx"):
    global DATAFRAMES_TEAMS

    # Caminho para um dos arquivos de equipe (ajuste conforme necessário)
    caminho_arquivo = f"{base_dir}/{escola}/Professores/{arquivo}"

    # Ler o arquivo Excel, pulando as primeiras 6 linhas
    df_team = pd.read_excel(caminho_arquivo, engine="openpyxl", skiprows=6)
    df_teams = df_team[["CPF", "Nome", "Função", "Telefone", "Email", "Observação"]].copy()

    # Atualiza o DATAFRAMES_TEAMS para ter todos os estudantes
    if DATAFRAMES_TEAMS is None or DATAFRAMES_TEAMS.empty:
        DATAFRAMES_TEAMS = df_teams.copy()

        # Criar o arquivo do zero caso ainda não exista
        df_teams.to_excel(output_file, sheet_name="Equipe", index=False, engine="openpyxl")

        print("Dataframe inicial criado")
    else:
        DATAFRAMES_TEAMS = pd.concat(
                [DATAFRAMES_TEAMS, df_teams], 
                ignore_index=True, 
                join="outer"  # Inclui todas as colunas existentes
            )
        
        if os.path.exists(output_file):
            # Carregar o arquivo Excel existente
            book = load_workbook(output_file)

            # Verifica se a aba "Estudantes" existe
            if "Equipe" not in book.sheetnames:
                print("Aba 'Equipe' não encontrada. Criando nova aba.")
                df_teams.to_excel(output_file, sheet_name="Equipe", index=False, engine="openpyxl")
                return

            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                sheet = book["Equipe"]
                startrow = sheet.max_row if sheet.max_row else 1  # Pega a última linha da aba
                df_teams.to_excel(writer, sheet_name="Equipe", index=False, startrow=startrow, header=False)

    total_linhas = len(DATAFRAMES_TEAMS)
    print(f"🧑‍🎓 Total de alunos atualizado: {total_linhas}")
    print(f"✅ Dados adicionados a {output_file}")
